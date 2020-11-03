# Jai Swaminarayan
import re
import os
import openpyxl
from datetime import datetime


def rollNoCellIsValid(rollNoCell):
    if(len(rollNoCell) == 2):
        rollNoCell = rollNoCell.upper()

        # Initializing
        checkAlphabets = True
        checkNumbers = False
        cellValid = False

        # First Index must be a letter
        if not(rollNoCell[0] >= 'A') and rollNoCell[0] <= 'Z':
            return None, None

        # Iterating entire string
        for i in range(1, len(rollNoCell)):
            if checkAlphabets and rollNoCell[i] >= 'A' and rollNoCell[i] <= 'Z':
                continue
            else:
                checkNumbers = True
                checkAlphabets = False

            # Checking Numbers at end
            if checkNumbers is True:
                if rollNoCell[i] >= '0' and rollNoCell[i] <= '9':
                    cellValid = True
                else:
                    cellValid = False

        # If Cell is true, split Char And Numbers
        # Ex: "A4" to 'A' and '4'
        if cellValid is True:
            temp = re.compile("([a-zA-Z]+)([0-9]+)")
            res = temp.match(rollNoCell).groups()

            # printing result
            col = ord(res[0]) - 64  # Ascii code of char
            row = int(res[1])

            return row, col

    # If length is not 2
    else:
        return None, None


def findRollNoCol():

    # Possible Text in files
    rollList = ['roll no', 'roll n', 'roll no.', 'roll',
                'rollno', 'rolln', 'rollno.',
                'roll number', 'roll num.', 'roll num',
                'rollnumber', 'rollnum.', 'rollnum']

    # Determining Max Row Search of Possible texts
    if max_row > 25:
        fetchTill = 25
    else:
        fetchTill = max_row

    print("Fetching Roll No in first ", fetchTill, " rows...", end=" ")

    # Iterating till rows Determined
    for i in range(1, fetchTill + 1):

        # Iterating over each column
        for j in range(1, max_col + 1):
            cell = mySheet.cell(row=i, column=j)

            # Reading value from object
            cell_value = cell.value
            if not(isinstance(cell_value, str)):
                continue

            # Converting to lower case
            cell_value = cell_value.lower()

            # Searching Possible texts as substring in all cols
            for item in rollList:
                if item in cell_value:
                    print("Done")
                    # Returning row and col where "roll no" was found
                    return i, j

    # if rollNoCell not found
    print("Done")
    return None, None


def insertData(rollNoRow, rollNoCol):
    # Getting current date
    date = datetime.now().strftime('%d')
    month = datetime.now().strftime('%b')
    year = datetime.now().strftime('%Y')

    # Adding full stop so that when last digit is 0,
    # it doesnt consider last session as 0
    full_date = month + "\n" + date + ",\n" + year + "."

    # # Reading .txt File
    file = open("list.txt", "r")

    # rollNoRow Storing to another variable
    tempRollNoRow = rollNoRow

    # Default row, col for "list.txt"
    row, col = 0, 0

    # By Default, considering that only one lecture was held
    multipleSessions = False

    # Getting date of last attendance recorded
    lastDateAt = mySheet.cell(row=rollNoRow, column=max_col)
    lastDateAt.alignment = openpyxl.styles.Alignment(wrapText=True)

    # Checking if attendance of same date exists
    if (lastDateAt.value is not None) and (full_date in lastDateAt.value):
        print("\nNOTE: More than 1 sessions found")
        print("So, naming column name as S-1, S-2, ..\n")

        # Multiple lectures were held
        multipleSessions = True

        # Getting last digit of last session
        if lastDateAt.value[-1] >= '0' and lastDateAt.value[-1] <= '9':
            lastSessionNo = int(lastDateAt.value[-1])
        else:
            # Renaming cellvalue to session1
            lastSessionNo = 1
            lastDateAt.value = full_date + "\n S-1"

    # Insert date at last column
    insertDateAt = mySheet.cell(row=rollNoRow, column=max_col+1)
    insertDateAt.alignment = openpyxl.styles.Alignment(wrapText=True,
                                                       horizontal='center')

    # If Multiple lectures were held
    if multipleSessions:
        insertDateAt.value = full_date + "\n S-" + str(lastSessionNo + 1)

    # if Multiple lectures were not held
    else:
        insertDateAt.value = full_date

    # Iterating over each attendee
    for attendee in file:

        # Remove blank lines
        if not attendee.strip():
            continue

        # Finding Roll no from entire String
        roll = re.findall("\d+", attendee)

        # Teachers wont have roll no, so ignoring them
        if not len(roll):
            print("Skipping", attendee, end="")
            continue

        # rollNoRow must be changed after each Iteration
        # so giving back the same value
        rollNoRow = tempRollNoRow

        # Iterating over all roll no of class
        for i in range(rollNoRow + 1, max_row + 1):

            # Getting cell_object of roll no of classes
            cell = mySheet.cell(row=i, column=rollNoCol)

            # NOTE: roll[0] is a `string`
            # If roll no of text file matches the roll no of csv file
            if(int(roll[0]) == cell.value):

                # Getting cell_object where new attendance will be recorded
                insertAtCell = mySheet.cell(row=i, column=max_col+1)

                # Store only first Number found from list
                insertAtCell.value = 'P'
                insertAtCell.alignment = openpyxl.styles.Alignment(
                    horizontal='center')

                # Printing those rollno who were recorded as present
                print(roll[0])
                break

            # Going to next record in .xls file
            rollNoRow += 1

        # Going to next record in .txt file
        row += 1


if __name__ == '__main__':
    filename = 'mca3-attendance.xlsx'
    mySheetName = 'sheet1'

    # If File does'nt exists
    if not(os.path.exists(filename)):
        print("File Does NOT Exists")
        exit(1)

    # open an Excel file and return a workbook
    workbook = openpyxl.load_workbook(filename)

    # By default: Sheet is not found
    sheetFound = False

    # Checking if specified sheet exists in workbook
    for sheet in workbook.sheetnames:
        if sheet.lower() == mySheetName:
            mySheetName = sheet
            sheetFound = True
            break

    # if sheetname not found
    if not sheetFound:
        exit(0)

    # Creating object of the sheet
    mySheet = workbook[mySheetName]

    # Finding max record and columns in .xls file
    max_row = mySheet.max_row
    max_col = mySheet.max_column

    # Trying to find roll no automatically
    rollNoRow, rollNoCol = findRollNoCol()

    # If roll no is not found automatically
    if rollNoCol is None:
        print("Cell not found !")
        while True:
            rollNoCell = input("Enter cell no of roll no"
                               "(Example: A4, B5)\n")
            rollNoRow, rollNoCol = rollNoCellIsValid(rollNoCell)
            if rollNoRow is not None:
                break

    # If roll no is found automatically
    else:
        print("Roll no found at [", rollNoRow, chr(rollNoCol + 64), "]")

    # Inserting attendance
    insertData(rollNoRow, rollNoCol)

    # Saving workbook
    workbook.save(filename)

    # Closing workbook
    workbook.close()
